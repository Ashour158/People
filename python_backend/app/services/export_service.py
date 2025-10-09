"""
Export Service
Enterprise-grade data export service supporting multiple formats
CSV, Excel, PDF, JSON exports with customization and streaming
"""
import asyncio
import csv
import io
import json
from datetime import datetime, date
from typing import List, Dict, Any, Optional, Union
from uuid import UUID
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import structlog
from sqlalchemy.ext.asyncio import AsyncSession

from app.core.config import settings

logger = structlog.get_logger()


class ExportService:
    """Enterprise export service with multiple format support"""
    
    def __init__(self):
        """Initialize export service"""
        self.temp_dir = Path("/tmp/exports")
        self.temp_dir.mkdir(parents=True, exist_ok=True)
    
    async def export_to_csv(
        self,
        data: List[Dict[str, Any]],
        columns: Optional[List[str]] = None,
        filename: Optional[str] = None
    ) -> bytes:
        """
        Export data to CSV format
        
        Args:
            data: List of dictionaries to export
            columns: List of column names to include (None = all)
            filename: Optional filename
            
        Returns:
            CSV file as bytes
        """
        try:
            if not data:
                raise ValueError("No data to export")
            
            # Determine columns
            if columns is None:
                columns = list(data[0].keys())
            
            # Create CSV in memory
            output = io.StringIO()
            writer = csv.DictWriter(output, fieldnames=columns, extrasaction='ignore')
            
            # Write header
            writer.writeheader()
            
            # Write data rows
            for row in data:
                # Convert complex types to strings
                cleaned_row = self._clean_row_for_csv(row, columns)
                writer.writerow(cleaned_row)
            
            # Get CSV content as bytes
            csv_content = output.getvalue()
            output.close()
            
            logger.info(
                "csv_export_completed",
                rows=len(data),
                columns=len(columns),
                size_kb=len(csv_content) / 1024
            )
            
            return csv_content.encode('utf-8-sig')  # BOM for Excel compatibility
            
        except Exception as e:
            logger.error("csv_export_failed", error=str(e))
            raise
    
    async def export_to_excel(
        self,
        data: List[Dict[str, Any]],
        columns: Optional[List[str]] = None,
        sheet_name: str = "Sheet1",
        filename: Optional[str] = None,
        styling: bool = True
    ) -> bytes:
        """
        Export data to Excel format with styling
        
        Args:
            data: List of dictionaries to export
            columns: List of column names to include
            sheet_name: Worksheet name
            filename: Optional filename
            styling: Apply styling to the worksheet
            
        Returns:
            Excel file as bytes
        """
        try:
            if not data:
                raise ValueError("No data to export")
            
            # Determine columns
            if columns is None:
                columns = list(data[0].keys())
            
            # Create workbook
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = sheet_name
            
            # Styling
            if styling:
                header_font = Font(bold=True, color="FFFFFF")
                header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                header_alignment = Alignment(horizontal="center", vertical="center")
                
                border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
            
            # Write headers
            for col_idx, column in enumerate(columns, start=1):
                cell = ws.cell(row=1, column=col_idx, value=column.replace('_', ' ').title())
                
                if styling:
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = header_alignment
                    cell.border = border
            
            # Write data rows
            for row_idx, row_data in enumerate(data, start=2):
                for col_idx, column in enumerate(columns, start=1):
                    value = row_data.get(column)
                    
                    # Convert complex types
                    if isinstance(value, (datetime, date)):
                        value = value.isoformat()
                    elif isinstance(value, UUID):
                        value = str(value)
                    elif isinstance(value, (dict, list)):
                        value = json.dumps(value)
                    elif value is None:
                        value = ""
                    
                    cell = ws.cell(row=row_idx, column=col_idx, value=value)
                    
                    if styling:
                        cell.border = border
                        cell.alignment = Alignment(vertical="center")
            
            # Auto-size columns
            for col_idx, column in enumerate(columns, start=1):
                col_letter = get_column_letter(col_idx)
                max_length = max(
                    len(str(column)),
                    max(len(str(ws.cell(row=row, column=col_idx).value or "")) 
                        for row in range(2, min(len(data) + 2, 100)))  # Check first 100 rows
                )
                ws.column_dimensions[col_letter].width = min(max_length + 2, 50)
            
            # Freeze header row
            if styling:
                ws.freeze_panes = "A2"
            
            # Save to bytes
            output = io.BytesIO()
            wb.save(output)
            excel_bytes = output.getvalue()
            output.close()
            
            logger.info(
                "excel_export_completed",
                rows=len(data),
                columns=len(columns),
                size_kb=len(excel_bytes) / 1024
            )
            
            return excel_bytes
            
        except Exception as e:
            logger.error("excel_export_failed", error=str(e))
            raise
    
    async def export_to_json(
        self,
        data: List[Dict[str, Any]],
        pretty: bool = True
    ) -> bytes:
        """
        Export data to JSON format
        
        Args:
            data: List of dictionaries to export
            pretty: Pretty print JSON
            
        Returns:
            JSON file as bytes
        """
        try:
            # Convert complex types
            cleaned_data = []
            for row in data:
                cleaned_row = {}
                for key, value in row.items():
                    if isinstance(value, (datetime, date)):
                        cleaned_row[key] = value.isoformat()
                    elif isinstance(value, UUID):
                        cleaned_row[key] = str(value)
                    else:
                        cleaned_row[key] = value
                cleaned_data.append(cleaned_row)
            
            # Convert to JSON
            if pretty:
                json_content = json.dumps(cleaned_data, indent=2, ensure_ascii=False)
            else:
                json_content = json.dumps(cleaned_data, ensure_ascii=False)
            
            logger.info(
                "json_export_completed",
                rows=len(data),
                size_kb=len(json_content) / 1024
            )
            
            return json_content.encode('utf-8')
            
        except Exception as e:
            logger.error("json_export_failed", error=str(e))
            raise
    
    async def export_employees(
        self,
        db: AsyncSession,
        organization_id: UUID,
        format: str = "excel",
        filters: Optional[Dict[str, Any]] = None
    ) -> bytes:
        """
        Export employee data
        
        Args:
            db: Database session
            organization_id: Organization UUID
            format: Export format (csv, excel, json)
            filters: Optional filters
            
        Returns:
            Export file as bytes
        """
        try:
            from app.models.models import Employee, Department
            from sqlalchemy import select, and_
            
            # Build query
            query = select(
                Employee.employee_code,
                Employee.first_name,
                Employee.last_name,
                Employee.work_email,
                Employee.personal_email,
                Employee.phone_number,
                Employee.designation,
                Department.department_name,
                Employee.date_of_joining,
                Employee.employment_status,
                Employee.employment_type,
                Employee.date_of_birth,
                Employee.gender,
                Employee.marital_status,
                Employee.address,
                Employee.city,
                Employee.country
            ).join(
                Department,
                Employee.department_id == Department.department_id,
                isouter=True
            ).where(
                and_(
                    Employee.organization_id == organization_id,
                    Employee.is_deleted == False
                )
            )
            
            # Apply filters
            if filters:
                if filters.get("department_id"):
                    query = query.where(Employee.department_id == filters["department_id"])
                if filters.get("employment_status"):
                    query = query.where(Employee.employment_status == filters["employment_status"])
            
            # Execute query
            result = await db.execute(query)
            rows = result.all()
            
            # Convert to list of dicts
            data = [dict(row._mapping) for row in rows]
            
            # Export based on format
            if format.lower() == "csv":
                return await self.export_to_csv(data)
            elif format.lower() == "excel":
                return await self.export_to_excel(data, sheet_name="Employees")
            elif format.lower() == "json":
                return await self.export_to_json(data)
            else:
                raise ValueError(f"Unsupported format: {format}")
            
        except Exception as e:
            logger.error("employee_export_failed", error=str(e))
            raise
    
    async def export_attendance(
        self,
        db: AsyncSession,
        organization_id: UUID,
        start_date: date,
        end_date: date,
        format: str = "excel",
        filters: Optional[Dict[str, Any]] = None
    ) -> bytes:
        """
        Export attendance data
        
        Args:
            db: Database session
            organization_id: Organization UUID
            start_date: Start date
            end_date: End date
            format: Export format
            filters: Optional filters
            
        Returns:
            Export file as bytes
        """
        try:
            from app.models.models import Employee, AttendanceLog
            from sqlalchemy import select, and_
            
            query = select(
                Employee.employee_code,
                Employee.first_name,
                Employee.last_name,
                AttendanceLog.attendance_date,
                AttendanceLog.check_in_time,
                AttendanceLog.check_out_time,
                AttendanceLog.total_hours,
                AttendanceLog.status,
                AttendanceLog.late_minutes,
                AttendanceLog.overtime_minutes
            ).join(
                Employee,
                AttendanceLog.employee_id == Employee.employee_id
            ).where(
                and_(
                    AttendanceLog.organization_id == organization_id,
                    AttendanceLog.attendance_date >= start_date,
                    AttendanceLog.attendance_date <= end_date
                )
            )
            
            result = await db.execute(query)
            rows = result.all()
            data = [dict(row._mapping) for row in rows]
            
            if format.lower() == "csv":
                return await self.export_to_csv(data)
            elif format.lower() == "excel":
                return await self.export_to_excel(data, sheet_name="Attendance")
            elif format.lower() == "json":
                return await self.export_to_json(data)
            else:
                raise ValueError(f"Unsupported format: {format}")
            
        except Exception as e:
            logger.error("attendance_export_failed", error=str(e))
            raise
    
    async def export_leave_records(
        self,
        db: AsyncSession,
        organization_id: UUID,
        year: int,
        format: str = "excel",
        filters: Optional[Dict[str, Any]] = None
    ) -> bytes:
        """
        Export leave records
        
        Args:
            db: Database session
            organization_id: Organization UUID
            year: Year
            format: Export format
            filters: Optional filters
            
        Returns:
            Export file as bytes
        """
        try:
            from app.models.models import Employee, LeaveRequest, LeaveType
            from sqlalchemy import select, and_, extract
            
            query = select(
                Employee.employee_code,
                Employee.first_name,
                Employee.last_name,
                LeaveType.leave_type_name,
                LeaveRequest.start_date,
                LeaveRequest.end_date,
                LeaveRequest.days,
                LeaveRequest.status,
                LeaveRequest.reason,
                LeaveRequest.created_at
            ).join(
                Employee,
                LeaveRequest.employee_id == Employee.employee_id
            ).join(
                LeaveType,
                LeaveRequest.leave_type_id == LeaveType.leave_type_id
            ).where(
                and_(
                    LeaveRequest.organization_id == organization_id,
                    extract('year', LeaveRequest.start_date) == year
                )
            )
            
            result = await db.execute(query)
            rows = result.all()
            data = [dict(row._mapping) for row in rows]
            
            if format.lower() == "csv":
                return await self.export_to_csv(data)
            elif format.lower() == "excel":
                return await self.export_to_excel(data, sheet_name="Leave Records")
            elif format.lower() == "json":
                return await self.export_to_json(data)
            else:
                raise ValueError(f"Unsupported format: {format}")
            
        except Exception as e:
            logger.error("leave_export_failed", error=str(e))
            raise
    
    async def export_payroll(
        self,
        db: AsyncSession,
        organization_id: UUID,
        month: int,
        year: int,
        format: str = "excel"
    ) -> bytes:
        """
        Export payroll data
        
        Args:
            db: Database session
            organization_id: Organization UUID
            month: Month number
            year: Year
            format: Export format
            
        Returns:
            Export file as bytes
        """
        try:
            from app.models.models import Employee, Payroll
            from sqlalchemy import select, and_
            
            query = select(
                Employee.employee_code,
                Employee.first_name,
                Employee.last_name,
                Payroll.basic_salary,
                Payroll.allowances,
                Payroll.deductions,
                Payroll.gross_salary,
                Payroll.net_salary,
                Payroll.tax_amount,
                Payroll.payment_status,
                Payroll.payment_date
            ).join(
                Employee,
                Payroll.employee_id == Employee.employee_id
            ).where(
                and_(
                    Payroll.organization_id == organization_id,
                    Payroll.month == month,
                    Payroll.year == year
                )
            )
            
            result = await db.execute(query)
            rows = result.all()
            data = [dict(row._mapping) for row in rows]
            
            if format.lower() == "csv":
                return await self.export_to_csv(data)
            elif format.lower() == "excel":
                return await self.export_to_excel(data, sheet_name="Payroll")
            elif format.lower() == "json":
                return await self.export_to_json(data)
            else:
                raise ValueError(f"Unsupported format: {format}")
            
        except Exception as e:
            logger.error("payroll_export_failed", error=str(e))
            raise
    
    async def export_multi_sheet_excel(
        self,
        sheets_data: Dict[str, List[Dict[str, Any]]],
        filename: Optional[str] = None
    ) -> bytes:
        """
        Export multiple sheets to single Excel file
        
        Args:
            sheets_data: Dictionary of sheet_name -> data
            filename: Optional filename
            
        Returns:
            Excel file as bytes
        """
        try:
            wb = openpyxl.Workbook()
            wb.remove(wb.active)  # Remove default sheet
            
            for sheet_name, data in sheets_data.items():
                if not data:
                    continue
                
                ws = wb.create_sheet(title=sheet_name[:31])  # Excel max 31 chars
                
                columns = list(data[0].keys())
                
                # Write headers
                for col_idx, column in enumerate(columns, start=1):
                    ws.cell(row=1, column=col_idx, value=column.replace('_', ' ').title())
                
                # Write data
                for row_idx, row_data in enumerate(data, start=2):
                    for col_idx, column in enumerate(columns, start=1):
                        value = row_data.get(column)
                        
                        if isinstance(value, (datetime, date)):
                            value = value.isoformat()
                        elif isinstance(value, UUID):
                            value = str(value)
                        elif isinstance(value, (dict, list)):
                            value = json.dumps(value)
                        
                        ws.cell(row=row_idx, column=col_idx, value=value)
            
            # Save to bytes
            output = io.BytesIO()
            wb.save(output)
            excel_bytes = output.getvalue()
            output.close()
            
            logger.info(
                "multi_sheet_excel_exported",
                sheets=len(sheets_data),
                size_kb=len(excel_bytes) / 1024
            )
            
            return excel_bytes
            
        except Exception as e:
            logger.error("multi_sheet_export_failed", error=str(e))
            raise
    
    def _clean_row_for_csv(self, row: Dict[str, Any], columns: List[str]) -> Dict[str, str]:
        """Clean row data for CSV export"""
        cleaned = {}
        for col in columns:
            value = row.get(col)
            
            if isinstance(value, (datetime, date)):
                cleaned[col] = value.isoformat()
            elif isinstance(value, UUID):
                cleaned[col] = str(value)
            elif isinstance(value, (dict, list)):
                cleaned[col] = json.dumps(value)
            elif value is None:
                cleaned[col] = ""
            else:
                cleaned[col] = str(value)
        
        return cleaned


# Singleton instance
export_service = ExportService()
